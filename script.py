# --- Standard Library Imports ------------------------------------------------

import argparse                       
import csv                       
import json                          
import logging                        
import os                          
import random                          
import threading                   
import re                              
import shutil                          
import sys                      
import time                           
from concurrent.futures import ThreadPoolExecutor, as_completed  
from datetime import datetime       
from urllib.parse import urlparse    
from pathlib import Path              
import openpyxl                     
import requests                    
from tqdm import tqdm             


# =============================================================================
# SECTION 1: GLOBAL CONFIGURATION
# All settings readable from environment variables.
# =============================================================================

DEFAULT_INPUT_FILE  = os.environ.get("VIMEO_INPUT_FILE","list.csv")
DEFAULT_OUTPUT_DIR  = os.environ.get("VIMEO_OUTPUT_DIR","./downloaded_videos")
LOG_DIR             = os.environ.get("VIMEO_LOG_DIR","logs")
PROGRESS_FILE       = os.environ.get("VIMEO_PROGRESS_FILE","progress.json")

MAX_RETRIES         = int(os.environ.get("VIMEO_MAX_RETRIES","5"))
RETRY_DELAY_BASE    = int(os.environ.get("VIMEO_RETRY_DELAY","5"))
RETRY_DELAY_429     = int(os.environ.get("VIMEO_RETRY_DELAY_429","60"))
CHUNK_SIZE          = int(os.environ.get("VIMEO_CHUNK_SIZE",str(1024 * 1024)))
REQUEST_TIMEOUT     = int(os.environ.get("VIMEO_TIMEOUT","60"))

BW_CHECK_INTERVAL   = int(os.environ.get("VIMEO_BW_INTERVAL",    "300"))   
BW_TEST_URL         = os.environ.get(
    "VIMEO_BW_TEST_URL",
    "http://speedtest.ftp.otenet.gr/files/test1Mb.db"                      
)
MAX_FILENAME_LEN    = 180                                               

# Bandwidth → thread count thresholds (Mbps)
# Format: (min_mbps, max_mbps, threads)
BANDWIDTH_TIERS = [
    (200, float("inf"), 10),   # > 200 Mbps  → 10 threads
    (100, 200,           6),   # 100–200 Mbps →  6 threads
    ( 50, 100,           4),   #  50–100 Mbps →  4 threads
    ( 10,  50,           2),   #  10–50 Mbps  →  2 threads
    (  0,  10,           1),   # < 10 Mbps    →  1 thread (serial)
]


# =============================================================================
# SECTION 2: HTTP STATUS CODE CLASSIFICATION
# =============================================================================

NO_RETRY_CODES = {
    400,   # Bad Request          
    401,   # Unauthorized       
    403,   # Forbidden           
    404,   # Not Found           
    405,   # Method Not Allowed  
    406,   # Not Acceptable      
    407,   # Proxy Auth Required  
    409,   # Conflict            
    410,   # Gone               
    411,   # Length Required     
    412,   # Precondition Failed  
    413,   # Payload Too Large    
    414,   # URI Too Long        
    415,   # Unsupported Media   
    416,   # Range Not Satisfiable 
    451,   # Unavailable Legal    

}

RETRY_WITH_WAIT_CODES = {

    429,   # Too Many Requests   
    503,   # Service Unavailable  

}

RETRYABLE_CODES = {

    500,   # Internal Server Error  
    502,   # Bad Gateway           
    504,   # Gateway Timeout         
    507,   # Insufficient Storage  
    508,   # Loop Detected         
    509,   # Bandwidth Exceeded      
    520,   # Unknown Error         
    521,   # Web Server Down         
    522,   # Connection Timed Out    
    523,   # Origin Unreachable   
    524,   # Timeout                 
    525,   # SSL Handshake Failed   
    526,   # Invalid SSL Cert      
    527,   # Railgun Error       
    530,   # Site Frozen             

}

# =============================================================================
# SECTION 3: LOGGING CONFIGURATION
# Each run creates a new dated log file inside the logs/ folder.
# A separate failed.log is maintained with structured failure reports.
# =============================================================================

def setup_logging(log_dir: str) -> tuple[logging.Logger, str]:
    Path(log_dir).mkdir(parents=True, exist_ok=True)        

    timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")   
    log_file  = str(Path(log_dir) / f"{timestamp}.log")        

    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s  %(levelname)-8s  %(message)s",
        handlers=[
            logging.FileHandler(log_file, encoding="utf-8"),   
            logging.StreamHandler(sys.stdout),              
        ],
    )

    log = logging.getLogger(__name__)
    log.info("Log file : %s", log_file)
    return log, log_file

def log_failed(
    log_dir:     str,
    video_id:    str,
    title:       str,
    url:         str,
    reason:      str,
    method:      str = "unknown",
    attempts:    int = 0,
    http_status: int = 0,
    file_size:   int = 0,
):
    failed_log   = Path(log_dir) / "failed.log"
    reason_lower = reason.lower()
    ts           = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    sep          = "=" * 80
    dash         = "-" * 80

    HTTP_MESSAGES = {
        400: ("Bad Request",           "Fix the URL — malformed or invalid parameters.",         False),
        401: ("Unauthorized",          "Re-export download URL from Vimeo — token expired.",     False),
        403: ("Forbidden",             "Video is private or access denied. Check permissions.",  False),
        404: ("Not Found",             "Video deleted or link expired. Re-export URL.",           False),
        405: ("Method Not Allowed",    "Wrong HTTP method. Check URL format.",                    False),
        406: ("Not Acceptable",        "Server cannot return requested format.",                  False),
        407: ("Proxy Auth Required",   "Configure proxy credentials.",                            False),
        409: ("Conflict",              "Resource state conflict. Re-export URL.",                 False),
        410: ("Gone",                  "Permanently deleted. Remove from input file.",            False),
        411: ("Length Required",       "Missing Content-Length header.",                          False),
        412: ("Precondition Failed",   "Server precondition not met.",                            False),
        413: ("Payload Too Large",     "Request too large.",                                      False),
        414: ("URI Too Long",          "URL too long for server.",                                False),
        415: ("Unsupported Media",     "Media type not supported.",                               False),
        416: ("Range Not Satisfiable", "Invalid byte range.",                                     False),
        429: ("Too Many Requests",     "Rate limited. Retry: --retry-failed",                     True),
        451: ("Unavailable Legal",     "DMCA or legal block. Cannot download.",                   False),
        500: ("Internal Server Error", "Server crashed. Retry: --retry-failed",                   True),
        502: ("Bad Gateway",           "Upstream error. Retry: --retry-failed",                   True),
        503: ("Service Unavailable",   "Server overloaded. Retry: --retry-failed",                True),
        504: ("Gateway Timeout",       "Upstream timeout. Retry: --retry-failed",                 True),
        520: ("Cloudflare Error",      "Origin unexpected response. Retry: --retry-failed",       True),
        521: ("Web Server Down",       "Origin offline. Retry: --retry-failed",                   True),
        522: ("Connection Timed Out",  "Origin timed out. Retry: --retry-failed",                 True),
        523: ("Origin Unreachable",    "Cannot route to origin. Retry: --retry-failed",           True),
        524: ("Cloudflare Timeout",    "Connection timed out. Retry: --retry-failed",             True),
        525: ("SSL Handshake Failed",  "SSL error. Retry: --retry-failed",                        True),
        526: ("Invalid SSL Cert",      "Bad SSL cert at origin. Retry: --retry-failed",           True),
    }

    if http_status and http_status in HTTP_MESSAGES:
        meaning, action, retryable = HTTP_MESSAGES[http_status]
        retry_label = "YES" if retryable else "NO"
        report = (
            f"{sep}\n"
            f"FAILED  |  {ts}\n"
            f"  ID      : {video_id}\n"
            f"  Title   : {title}\n"
            f"  Method  : {method}  |  Attempts: {attempts}/{MAX_RETRIES}\n"
            f"  HTTP    : {http_status} {meaning}\n"
            f"  Retry   : {retry_label}\n"
            f"  Action  : {action}\n"
            f"  URL     : {url}\n"
            f"{sep}\n\n"
        )

    elif http_status and http_status >= 400:
        retryable   = http_status >= 500
        retry_label = "YES" if retryable else "NO"
        report = (
            f"{sep}\n"
            f"FAILED  |  {ts}\n"
            f"  ID      : {video_id}\n"
            f"  Title   : {title}\n"
            f"  Method  : {method}  |  Attempts: {attempts}/{MAX_RETRIES}\n"
            f"  HTTP    : {http_status} (unrecognised)\n"
            f"  Retry   : {retry_label}\n"
            f"  Action  : {'Retry: --retry-failed' if retryable else 'Fix URL or remove entry'}\n"
            f"  URL     : {url}\n"
            f"{sep}\n\n"
        )

    else:
        if "timeout" in reason_lower:
            category    = "TIMEOUT"
            explanation = f"No response within {REQUEST_TIMEOUT}s. Check connection speed or increase VIMEO_TIMEOUT."
            retryable   = True
            action      = "Retry: python download_videos.py --retry-failed\n   Or increase: set VIMEO_TIMEOUT=120"

        elif "connection" in reason_lower or "network" in reason_lower:
            category    = "NETWORK ERROR"
            explanation = "DNS failure, no internet, firewall blocking, or server unreachable."
            retryable   = True
            action      = "Check internet connection, then retry: python download_videos.py --retry-failed"

        elif "incomplete" in reason_lower:
            sz          = f"{file_size:,} bytes ({file_size/1e6:.2f} MB)" if file_size else "unknown"
            category    = "INCOMPLETE DOWNLOAD"
            explanation = f"Only {sz} received before connection dropped. Network interruption mid-download."
            retryable   = True
            action      = "Retry: python download_videos.py --retry-failed"

        elif "private" in reason_lower or "members only" in reason_lower:
            category    = "PRIVATE VIDEO"
            explanation = "yt-dlp: video is private or members-only. Account lacks download permission."
            retryable   = False
            action      = "Check Vimeo privacy settings. Generate a direct download link if you own the video."

        elif "geo" in reason_lower or "not available in your country" in reason_lower:
            category    = "GEO-BLOCKED"
            explanation = "Video restricted in current geographic region."
            retryable   = False
            action      = "Use a VPN set to a permitted region, or check Vimeo geo-restriction settings."

        elif "no info" in reason_lower or "returned no info" in reason_lower:
            category    = "NO VIDEO INFO"
            explanation = "yt-dlp could not extract info — video may be deleted or URL format unsupported."
            retryable   = False
            action      = "Verify URL is valid in browser. Re-export download URL from Vimeo."

        elif "not installed" in reason_lower:
            category    = "MISSING LIBRARY"
            explanation = "yt-dlp is not installed on this system."
            retryable   = False
            action      = "Run: pip install yt-dlp"

        else:
            category    = "UNKNOWN ERROR"
            explanation = "Unexpected error — check the dated log file in logs/ for the full traceback."
            retryable   = True
            action      = "Check logs/ for traceback. Retry: python download_videos.py --retry-failed"

        retry_label = "YES" if retryable else "NO"
        report = (
            f"{sep}\n"
            f"FAILED  |  {ts}\n"
            f"  ID          : {video_id}\n"
            f"  Title       : {title}\n"
            f"  Method      : {method}  |  Attempts: {attempts}/{MAX_RETRIES}\n"
            f"  Category    : {category}\n"
            f"  Retryable   : {retry_label}\n"
            f"{dash}\n"
            f"  Explanation : {explanation}\n"
            f"  Raw Error   : {reason}\n"
            f"  Action      : {action}\n"
            f"{dash}\n"
            f"  URL         : {url}\n"
            f"{sep}\n\n"
        )

    with _failed_lock:                                      
        with open(failed_log, "a", encoding="utf-8") as f:
            f.write(report)

# =============================================================================
# SECTION 4: Progress Tracking (progress.json)
# Track every Video ID with a status so the script can resume
# =============================================================================

def load_progress(progress_file: str) -> dict:
    path = Path(progress_file)
    if not path.exists():                                
        return {}

    try:
        with open(path, encoding="utf-8") as f:
            data = json.load(f)                               
        if not isinstance(data, dict):                       
            raise ValueError(f"Expected dict, got {type(data).__name__}")
        return data                                          

    except json.JSONDecodeError as exc:                      
        corrupt_path = path.with_suffix(".corrupt")           
        path.rename(corrupt_path)                           
        logging.getLogger(__name__).warning(
            "progress.json is corrupt (%s) — renamed to %s — starting fresh",
            exc, corrupt_path.name,
        )
        return {}                                             

    except (OSError, ValueError) as exc:                     
        logging.getLogger(__name__).warning(
            "Could not read progress.json (%s) — starting fresh", exc
        )
        return {}


_progress_lock = threading.Lock()                             
_failed_lock   = threading.Lock()                              


def save_progress(progress_file: str, progress: dict):
    path     = Path(progress_file)
    tmp_path = path.with_suffix(".tmp")                     

    with _progress_lock:                                      
        with open(tmp_path, "w", encoding="utf-8") as f:
            json.dump(progress, f, indent=2, ensure_ascii=False)
        tmp_path.replace(path)                              


def init_progress(rows: list[dict], progress: dict) -> dict:
    for row in rows:
        vid_id = str(row.get("video_id") or "").strip()
        if vid_id and vid_id not in progress:
            progress[vid_id] = {
                "status": "pending",
                "title":  str(row.get("title") or ""),
                "url":    str(row.get("download_link") or ""),
            }
    return progress


# =============================================================================
# SECTION 5a: HELPER — sanitize_filename()
# Strips characters illegal in Windows/macOS/Linux filenames.
# =============================================================================

def sanitize_filename(name: str) -> str:
    name = re.sub(r'[\\/*?:"<>|]', "_", name)               
    name = name.strip().strip(".")                           
    return name or "untitled"

# =============================================================================
# SECTION 5b: HELPER — build_filename()
# Builds output filename as: <VideoID>_<VideoTitle>.<ext>
# =============================================================================

def build_filename(vid_id: str, title: str, ext: str = "") -> str:
    safe_id    = sanitize_filename(vid_id) if vid_id else "no_id"  
    safe_title = sanitize_filename(title)                          

    safe_title = re.sub(r"\s+", "_", safe_title)                   
    safe_title = re.sub(r"_+", "_", safe_title)                  

    if safe_title and safe_title != "untitled":
        stem = f"{safe_id}_{safe_title}"                            
    else:
        stem = safe_id                                            

    max_stem = MAX_FILENAME_LEN - len(ext)                         
    if len(stem) > max_stem:
        stem = stem[:max_stem].rstrip("_")                      

    return f"{stem}{ext}"                                          

# =============================================================================
# SECTION 6: CORE — read_input_file()
# Reads .xlsx or .csv and returns list of video record dicts.
# =============================================================================

def read_input_file(path: str) -> list[dict]:
    path = Path(path)
    rows = []

    if path.suffix.lower() in (".xlsx", ".xls"):
        wb      = openpyxl.load_workbook(path)
        ws      = wb.active
        headers = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]
        for row in ws.iter_rows(min_row=2, values_only=True):
            rd = dict(zip(headers, row))
            if not any(rd.values()):
                continue
            rows.append({
                "video_id":      rd.get("video id", ""),
                "title":         rd.get("video title", ""),
                "download_link": rd.get("download url", ""),
            })

    elif path.suffix.lower() == ".csv":
        with open(path, newline="", encoding="utf-8-sig") as f:
            reader = csv.DictReader(f)
            for row in reader:
                norm = {k.strip().lower(): v for k, v in row.items()}
                if not any(norm.values()):
                    continue
                rows.append({
                    "video_id":      norm.get("video id", ""),
                    "title":         norm.get("video title", ""),
                    "download_link": norm.get("download url", ""),
                })

    else:
        logging.getLogger(__name__).error(
            "Unsupported format '%s' — use .csv or .xlsx", path.suffix
        )
        sys.exit(1)

    return rows


# =============================================================================
# SECTION 7a: HELPER — is_valid_url()
# Validates URL format before attempting download.
# =============================================================================

def is_valid_url(url: str) -> bool:
    if not url or not isinstance(url, str):                   
        return False
    try:
        result = urlparse(url.strip())                        
        return (                                               
            result.scheme in ("http", "https")                 
            and bool(result.netloc)                          
            and len(url) < 2048                               
        )
    except Exception:
        return False                                           

# =============================================================================
# SECTION 7b: HELPER — is_direct_download()
# Detects whether a URL is a direct file link (requests) or
# a watch-page link (yt-dlp).
# =============================================================================

def is_direct_download(url: str) -> bool:
    direct_signals = (
        "progressive_redirect",        
        "player.vimeo.com/external",   
        ".mp4", ".mov", ".webm",     
        ".mkv", ".avi",
    )
    return any(s in url.lower() for s in direct_signals)

# =============================================================================
# SECTION 8a: BANDWIDTH MONITOR — BandwidthMonitor
# Measures current internet bandwidth by timing a small test download.
# =============================================================================

class BandwidthMonitor:
    def __init__(self, log: logging.Logger):
        self.log            = log
        self.current_mbps   = 0.0                              
        self.current_threads = 1                         
        self.last_check     = 0.0                             
        self.lock           = threading.Lock()                 

    def measure_bandwidth(self) -> float:
        try:
            start    = time.time()                            
            response = requests.get(
                BW_TEST_URL,
                timeout=(5, 30),                              
                stream=True,
            )
            response.raise_for_status()

            bytes_received = 0
            for chunk in response.iter_content(chunk_size=65536): 
                if chunk:
                    bytes_received += len(chunk)

            elapsed = time.time() - start                   
            if elapsed <= 0 or bytes_received == 0:
                return 0.0

            mbps = (bytes_received * 8) / (elapsed * 1_000_000)  
            return round(mbps, 2)

        except Exception as exc:
            self.log.warning("Bandwidth test failed: %s — keeping current threads", exc)
            return self.current_mbps                          

    def get_threads_for_bandwidth(self, mbps: float) -> int:
        for min_mbps, max_mbps, threads in BANDWIDTH_TIERS:
            if min_mbps <= mbps < max_mbps:
                return threads
        return 1                                             

    def check(self, force: bool = False) -> tuple[float, int, bool]:
        now = time.time()
        if not force and (now - self.last_check) < BW_CHECK_INTERVAL:
            return self.current_mbps, self.current_threads, False

        with self.lock:
            self.log.info("-" * 60)
            self.log.info("BANDWIDTH CHECK — measuring...")

            mbps           = self.measure_bandwidth()
            new_threads    = self.get_threads_for_bandwidth(mbps)
            changed        = new_threads != self.current_threads
            old_threads    = self.current_threads

            self.current_mbps    = mbps
            self.current_threads = new_threads
            self.last_check      = time.time()

            tier_label = next(
                (f">{t[0]} Mbps" if t[1] == float("inf") else f"{t[0]}–{t[1]} Mbps"
                 for t in BANDWIDTH_TIERS if t[0] <= mbps < t[1]),
                "unknown"
            )
            self.log.info(
                "BANDWIDTH — %.1f Mbps (%s) → %d thread(s)",
                mbps, tier_label, new_threads,
            )
            if changed:
                self.log.info(
                    "THREADS CHANGED — %d → %d (bandwidth: %.1f Mbps)",
                    old_threads, new_threads, mbps,
                )
            else:
                self.log.info(
                    "THREADS UNCHANGED — staying at %d (bandwidth: %.1f Mbps)",
                    new_threads, mbps,
                )

            self.log.info("-" * 60)
            return mbps, new_threads, changed


# =============================================================================
# SECTION 8b: DOWNLOAD — download_with_ytdlp()
# Downloads Vimeo/YouTube watch-page URLs via yt-dlp.
# =============================================================================

def download_with_ytdlp(
    url:         str,
    output_path: Path,
    log:         logging.Logger,
) -> tuple[bool, str]:
    try:
        import yt_dlp                                        
    except ImportError:
        log.error("yt-dlp not installed — run: pip install yt-dlp")
        return False, "yt-dlp not installed"

    ydl_opts = {
        "outtmpl":      str(output_path.with_suffix("")) + ".%(ext)s",
        "quiet":        False,
        "no_warnings":  False,
        "ignoreerrors": False,
        "retries":      MAX_RETRIES,
        "noprogress":   False,
    }

    try:
        log.info("  yt-dlp downloading: %s", url)
        with yt_dlp.YoutubeDL(ydl_opts) as ydl:
            info = ydl.extract_info(url, download=True)
            if info is None:
                return False, "yt-dlp returned no info — URL may be private or deleted"
        return True, "ok"

    except yt_dlp.utils.DownloadError as exc:                 
        reason = f"yt-dlp DownloadError: {exc}"
        log.error("  %s", reason)
        return False, reason

    except yt_dlp.utils.ExtractorError as exc:                 
        reason = f"yt-dlp ExtractorError: {exc}"
        log.error("  %s", reason)
        return False, reason

    except yt_dlp.utils.PostProcessingError as exc:           
        reason = f"yt-dlp PostProcessingError: {exc}"
        log.error("  %s", reason)
        return False, reason

    except KeyboardInterrupt:                                
        raise                                                  

    except Exception as exc:                                 
        reason = f"yt-dlp unexpected error: {exc}"
        log.error("  %s", reason)
        log.exception("  Full traceback:")                     
        return False, reason


# =============================================================================
# SECTION 9: DOWNLOAD — download_direct()
# Called per-thread by the parallel download engine.
# =============================================================================

def download_direct(
    url:         str,
    output_path: Path,
    log:         logging.Logger,
) -> tuple[bool, str]:
    reason = "Download did not start"                        

    for attempt in range(1, MAX_RETRIES + 1):

        try:
            with requests.get(
                url,
                stream=True,
                timeout=(10, REQUEST_TIMEOUT),             
            ) as r:
                if r.status_code in NO_RETRY_CODES:
                    reason = (
                        f"HTTP {r.status_code} — permanent failure, will not retry. "
                        f"Fix the URL or permissions and re-run."
                    )
                    log.error("  %s", reason)
                    return False, reason                      
                if r.status_code in RETRY_WITH_WAIT_CODES:
                    reason = f"HTTP {r.status_code} — rate limited or temporarily unavailable"
                    log.warning(
                        "  HTTP %d — waiting %ds before retry (attempt %d/%d)",
                        r.status_code, RETRY_DELAY_429, attempt, MAX_RETRIES,
                    )
                    time.sleep(RETRY_DELAY_429)              
                    continue                         
                if r.status_code in RETRYABLE_CODES:
                    reason = (
                        f"HTTP {r.status_code} — transient server error, "
                        f"retrying (attempt {attempt}/{MAX_RETRIES})"
                    )
                    log.warning("  %s", reason)
                    output_path.unlink(missing_ok=True)
                    if attempt < MAX_RETRIES:
                        delay = (RETRY_DELAY_BASE * attempt) + random.uniform(0, 2) 
                        log.warning("  Retrying in %.1fs...", delay)
                        time.sleep(delay)
                    continue
                if 400 <= r.status_code < 500:
                    reason = (
                        f"HTTP {r.status_code} — unrecognised client error, "
                        f"treating as permanent failure"
                    )
                    log.error("  %s", reason)
                    return False, reason

                r.raise_for_status()

                total = int(r.headers.get("Content-Length", 0))

                if total > 0:
                    free_bytes = shutil.disk_usage(output_path.parent).free
                    if free_bytes < total:
                        reason = (
                            f"Insufficient disk space — "
                            f"need {total/1e6:.1f} MB, "
                            f"only {free_bytes/1e6:.1f} MB free"
                        )
                        log.error("  DISK SPACE: %s", reason)
                        return False, reason               
                with (
                    open(output_path, "wb") as f,
                    tqdm(
                        total=total,
                        unit="B",
                        unit_scale=True,
                        unit_divisor=1024,
                        desc=f"  {output_path.name[:35]}",
                        leave=False,
                        ncols=90,
                    ) as bar,
                ):
                    downloaded = 0
                    for chunk in r.iter_content(chunk_size=CHUNK_SIZE):
                        if chunk:
                            f.write(chunk)
                            bar.update(len(chunk))
                            downloaded += len(chunk)

                if total > 0 and downloaded < total:
                    reason = f"Incomplete download: received {downloaded:,} of {total:,} bytes"
                    log.warning("  %s — retrying", reason)
                    output_path.unlink(missing_ok=True)
                    delay = (RETRY_DELAY_BASE * attempt) + random.uniform(0, 2)  
                    time.sleep(delay)
                    continue

                actual_size = output_path.stat().st_size if output_path.exists() else 0
                if actual_size < 1024:                        
                    reason = (
                        f"File integrity failed — "
                        f"only {actual_size} bytes on disk after download"
                    )
                    log.warning("  %s — retrying", reason)
                    output_path.unlink(missing_ok=True)
                    delay = (RETRY_DELAY_BASE * attempt) + random.uniform(0, 2)  
                    time.sleep(delay)
                    continue

                return True, "ok"                              

        except requests.exceptions.Timeout:
            reason = f"Timeout — no response within {REQUEST_TIMEOUT}s"
            log.warning("  Attempt %d/%d — %s", attempt, MAX_RETRIES, reason)

        except requests.exceptions.SSLError as exc:            
            reason = f"SSL error: {exc}"
            log.warning("  Attempt %d/%d — %s", attempt, MAX_RETRIES, reason)

        except requests.exceptions.ConnectionError as exc:
            reason = f"Connection error: {exc}"
            log.warning("  Attempt %d/%d — %s", attempt, MAX_RETRIES, reason)

        except requests.exceptions.HTTPError as exc:
            reason = f"HTTP error: {exc}"
            log.warning("  Attempt %d/%d — %s", attempt, MAX_RETRIES, reason)

        except OSError as exc:                                
            reason = f"File system error: {exc}"
            log.error("  Attempt %d/%d — %s", attempt, MAX_RETRIES, reason)
            output_path.unlink(missing_ok=True)
            return False, reason                               

        except Exception as exc:                               
            reason = f"Unexpected error: {exc}"
            log.error("  Attempt %d/%d — %s", attempt, MAX_RETRIES, reason)
            log.exception("  Full traceback:")                 

        output_path.unlink(missing_ok=True)                 

        if attempt < MAX_RETRIES:
            delay = (RETRY_DELAY_BASE * attempt) + random.uniform(0, 2)
            log.warning("  Retrying in %.1fs... (%d/%d)", delay, attempt, MAX_RETRIES)
            time.sleep(delay)

    return False, reason                                   

# =============================================================================
# SECTION 10: ORCHESTRATOR — download_video()
# Full per-video download flow. Uses Video ID as primary key filename.
# Updates progress.json after every video.
# =============================================================================

def download_video(
    row:           dict,
    output_dir:    Path,
    progress:      dict,
    progress_file: str,
    log_dir:       str,
    log:           logging.Logger,
) -> str:
    url    = str(row.get("download_link") or "").strip()
    title  = str(row.get("title")         or "").strip()
    vid_id = str(row.get("video_id")      or "").strip()

    if not vid_id:
        log.warning("NO VIDEO ID — title='%s' — using 'no_id'", title)
        vid_id = "no_id"

    if progress.get(vid_id, {}).get("status") == "ok":
        log.info("SKIP  [ID=%s] already completed in previous run", vid_id)
        return "skipped"

    if not url:
        log.warning("NO URL  [ID=%s] title='%s' — skipping", vid_id, title)
        progress[vid_id] = {"status": "no_url", "title": title, "url": ""}
        save_progress(progress_file, progress)
        return "no_url"

    if not is_valid_url(url):
        reason = f"Invalid URL format — '{url}' is not a valid http/https URL"
        log.error("  INVALID URL [ID=%s]: %s", vid_id, url)
        progress[vid_id] = {"status": "failed", "title": title, "url": url, "reason": reason}
        save_progress(progress_file, progress)
        log_failed(log_dir=log_dir, video_id=vid_id, title=title,
                   url=url, reason=reason, method="none", attempts=0)
        return "failed"

    if is_direct_download(url):
        ext         = Path(url.split("?")[0]).suffix or ".mp4"  
        filename    = build_filename(vid_id, title, ext)        
        output_path = output_dir / filename
    else:
        filename    = build_filename(vid_id, title, "")          
        output_path = output_dir / filename

    safe_id  = sanitize_filename(vid_id) if vid_id else "no_id"
    existing = list(output_dir.glob(f"{safe_id}_*.*")) or list(output_dir.glob(f"{safe_id}.*"))
    if existing:
        log.info("SKIP  [ID=%s] already on disk: %s", vid_id, existing[0].name)
        progress[vid_id] = {"status": "ok", "title": title, "url": url}
        save_progress(progress_file, progress)
        return "skipped"

    log.info("START [ID=%s] %s", vid_id, title)
    log.info("  URL      : %s", url)
    log.info("  Saving as: %s", output_path.name)

    if is_direct_download(url):
        success, reason = download_direct(url, output_path, log)
    else:
        success, reason = download_with_ytdlp(url, output_path, log)

    if success:
        log.info("OK    [ID=%s] %s", vid_id, title)
        progress[vid_id] = {"status": "ok", "title": title, "url": url}
    else:
        log.error("FAIL  [ID=%s] %s — %s", vid_id, title, reason)
        progress[vid_id] = {"status": "failed", "title": title, "url": url, "reason": reason}

        http_match  = re.search(r"HTTP (\d{3})", reason)
        http_status = int(http_match.group(1)) if http_match else 0
        method      = "requests" if is_direct_download(url) else "yt-dlp"

        log_failed(
            log_dir     = log_dir,
            video_id    = vid_id,
            title       = title,
            url         = url,
            reason      = reason,
            method      = method,
            attempts    = MAX_RETRIES,
            http_status = http_status,
        )
        output_path.unlink(missing_ok=True)                    

    save_progress(progress_file, progress)
    return "ok" if success else "failed"


# =============================================================================
# SECTION 11: THREAD-SAFE OVERALL PROGRESS DISPLAY
# Persistent progress bar: files done/total, % complete, ETA.
# =============================================================================

class OverallProgress:
    def __init__(self, total: int, log: logging.Logger):
        self.total      = total
        self.done       = 0
        self.failed     = 0
        self.skipped    = 0
        self.start_time = time.time()
        self.log        = log
        self.lock       = threading.Lock()                     
        self.bar        = tqdm(
            total=total,
            unit="file",
            desc="Overall",
            ncols=90,
            position=0,
            leave=True,
            bar_format="{l_bar}{bar}| {n_fmt}/{total_fmt} [{elapsed}<{remaining}, {rate_fmt}]",
        )

    def update(self, result: str, elapsed_secs: float):
        with self.lock:                                  
            self.done += 1
            if result == "failed":
                self.failed += 1
            elif result in ("skipped", "no_url"):
                self.skipped += 1
            self.bar.update(1)

        elapsed_total = time.time() - self.start_time
        avg_per_file  = elapsed_total / self.done
        eta_secs      = avg_per_file * (self.total - self.done)
        pct           = (self.done / self.total) * 100

        self.log.info(
            "PROGRESS  %d/%d (%.1f%%) | Done:%d Skipped:%d Failed:%d | ETA:%s",
            self.done, self.total, pct,
            self.done - self.failed - self.skipped,
            self.skipped, self.failed,
            self._format_time(eta_secs),
        )

    def close(self):
        self.bar.close()

    @staticmethod
    def _format_time(seconds: float) -> str:
        s = int(seconds)
        if s >= 3600:
            return f"{s // 3600}h {(s % 3600) // 60}m"
        elif s >= 60:
            return f"{s // 60}m {s % 60}s"
        else:
            return f"{s}s"


# =============================================================================
# SECTION 12: ENTRY POINT 
# =============================================================================

def main():

    global LOG_DIR, PROGRESS_FILE, MAX_RETRIES, RETRY_DELAY_BASE
    global RETRY_DELAY_429, CHUNK_SIZE, REQUEST_TIMEOUT

    # --- Step 1: Parse CLI arguments -----------------------------------------

    _epilog = (
        "Environment variables (all overridden by CLI args):\n"
        "  VIMEO_INPUT_FILE       Input .csv or .xlsx file\n"
        "  VIMEO_OUTPUT_DIR       Output folder for downloaded videos\n"
        "  VIMEO_LOG_DIR          Folder for dated log files\n"
        "  VIMEO_PROGRESS_FILE    Path to progress.json\n"
        "  VIMEO_MAX_RETRIES      Max retry attempts (int)\n"
        "  VIMEO_RETRY_DELAY      Base retry delay in seconds (int)\n"
        "  VIMEO_RETRY_DELAY_429  Delay for HTTP 429/503 in seconds (int)\n"
        "  VIMEO_CHUNK_SIZE       Download chunk size in bytes (int)\n"
        "  VIMEO_TIMEOUT          HTTP read timeout in seconds (int)\n"
        "\n"
        "Examples:\n"
        "  python download_videos.py\n"
        "  python download_videos.py -i list.csv -o E:/Vimeo_Archive\n"
        "  python download_videos.py --retry-failed\n"
    )

    parser = argparse.ArgumentParser(
        description="Econ Engineering — Vimeo Video Bulk Downloader v3.0.0",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=_epilog,
    )
    parser.add_argument("--input",  "-i", default=DEFAULT_INPUT_FILE,
                        help=f"Input file [env: VIMEO_INPUT_FILE] (default: {DEFAULT_INPUT_FILE})")
    parser.add_argument("--output", "-o", default=DEFAULT_OUTPUT_DIR,
                        help=f"Output folder [env: VIMEO_OUTPUT_DIR] (default: {DEFAULT_OUTPUT_DIR})")
    parser.add_argument("--log-dir",       default=LOG_DIR,
                        help=f"Log folder [env: VIMEO_LOG_DIR] (default: {LOG_DIR})")
    parser.add_argument("--progress-file", default=PROGRESS_FILE,
                        help=f"Progress file [env: VIMEO_PROGRESS_FILE] (default: {PROGRESS_FILE})")
    parser.add_argument("--retry-failed",  action="store_true",
                        help="Re-attempt only videos marked as failed in progress.json")
    args = parser.parse_args()

    LOG_DIR       = args.log_dir
    PROGRESS_FILE = args.progress_file

    # --- Step 2: Logging + output dir ----------------------------------------

    log, log_file = setup_logging(args.log_dir)
    output_dir    = Path(args.output)
    output_dir.mkdir(parents=True, exist_ok=True)

    # --- Step 3: Startup banner ----------------------------------------------

    log.info("=" * 60)
    log.info("  Econ Engineering Video Downloader  v3.0.0")
    log.info("=" * 60)
    log.info("Input        : %s", args.input)
    log.info("Output       : %s", output_dir.resolve())
    log.info("Progress file: %s", PROGRESS_FILE)
    log.info("Log folder   : %s", Path(LOG_DIR).resolve())
    log.info("Python       : %s", sys.version.split()[0])
    log.info("Filename key : ID + Title (e.g. 1156438752_PWC_Delhi.mp4)")
    log.info("Concurrency  : Dynamic (bandwidth-based, checked every %ds)", BW_CHECK_INTERVAL)
    log.info("-" * 60)

    # --- Step 4: Library check -----------------------------------------------

    all_ok = True
    for lib, pkg in [("requests","requests"), ("openpyxl","openpyxl"),
                     ("tqdm","tqdm"), ("yt_dlp","yt-dlp")]:
        try:
            mod = __import__(lib)
            ver = getattr(mod, "__version__", None) or getattr(
                getattr(mod, "version", None), "__version__", "?")
            log.info("%-10s : OK  (v%s)", lib, ver)
        except ImportError:
            log.error("%-10s : MISSING — run: pip install %s", lib, pkg)
            all_ok = False

    if not all_ok:
        log.error("Run: pip install requests openpyxl yt-dlp tqdm")
        sys.exit(1)

    # --- Step 5: Input file check --------------------------------------------

    log.info("-" * 60)
    if not Path(args.input).exists():
        log.error("INPUT FILE NOT FOUND: '%s'", args.input)
        log.error("Make sure '%s' is in the same folder as this script.", args.input)
        sys.exit(1)

    # --- Step 6: Read input file ---------------------------------------------

    rows  = read_input_file(args.input)
    total = len(rows)
    log.info("Found %d video(s) in input file.", total)

    # --- Step 7: Load / init progress ----------------------------------------

    progress = load_progress(PROGRESS_FILE)
    progress = init_progress(rows, progress)

    reset_count = 0
    for vid_id, entry in progress.items():
        if entry.get("status") == "ok":
            safe_id = sanitize_filename(vid_id)
            if not list(output_dir.glob(f"{safe_id}.*")):
                log.warning(
                    "RESET [ID=%s] marked ok but file missing on disk — re-queuing", vid_id
                )
                entry["status"] = "pending"
                reset_count += 1

    if reset_count > 0:
        log.warning("%d video(s) reset — files missing on disk", reset_count)

    save_progress(PROGRESS_FILE, progress)

    done_count    = sum(1 for v in progress.values() if v["status"] == "ok")
    failed_count  = sum(1 for v in progress.values() if v["status"] == "failed")
    pending_count = sum(1 for v in progress.values() if v["status"] == "pending")
    log.info("State — Done:%d | Failed:%d | Pending:%d | Reset:%d",
             done_count, failed_count, pending_count, reset_count)

    # --- Step 8: Filter rows -------------------------------------------------

    if args.retry_failed:
        rows = [r for r in rows
                if progress.get(str(r.get("video_id","")).strip(),{}).get("status") == "failed"]
        log.info("--retry-failed: %d video(s) to re-attempt", len(rows))
    else:
        rows = [r for r in rows
                if progress.get(str(r.get("video_id","")).strip(),{}).get("status") != "ok"]
        log.info("Remaining: %d video(s)", len(rows))

    if not rows:
        log.info("Nothing to download — all videos already completed.")
        log.info("To retry failures: python download_videos.py --retry-failed")
        sys.exit(0)

    log.info("=" * 60)

    # --- Step 9: Download loop -----------------------------------------------

    counts   = {"ok": 0, "skipped": 0, "failed": 0, "no_url": 0}
    overall  = OverallProgress(len(rows), log)

    # --- Bandwidth monitor — initial check before starting loop --------------

    bw_monitor = BandwidthMonitor(log)
    log.info("=" * 60)
    log.info("INITIAL BANDWIDTH CHECK")
    mbps, n_threads, _ = bw_monitor.check(force=True)         # Force check immediately
    log.info("Starting with %d thread(s) at %.1f Mbps", n_threads, mbps)
    log.info("=" * 60)

    # --- Parallel download loop with dynamic concurrency ---------------------

    total_rows    = len(rows)
    completed     = 0                                         

    try:
        remaining = list(rows)                                

        while remaining:

            _, n_threads, changed = bw_monitor.check()
            batch_size = n_threads                          

            batch    = remaining[:batch_size]
            remaining = remaining[batch_size:]

            log.info(
                "BATCH — %d thread(s) | processing %d video(s) | %d remaining after this batch",
                n_threads, len(batch), len(remaining),
            )

            with ThreadPoolExecutor(max_workers=n_threads) as executor:
                future_to_row = {
                    executor.submit(
                        download_video,
                        row,
                        output_dir,
                        progress,
                        PROGRESS_FILE,
                        LOG_DIR,
                        log,
                    ): row
                    for row in batch
                }

                for future in as_completed(future_to_row):
                    row = future_to_row[future]
                    t_start = time.time()
                    try:
                        result = future.result()               
                    except Exception as exc:
                        vid_id = str(row.get("video_id","")).strip()
                        log.error("Thread error for [ID=%s]: %s", vid_id, exc)
                        result = "failed"

                    counts[result] += 1
                    completed += 1
                    overall.update(result, time.time() - t_start)
                    log.info(
                        "OVERALL  %d/%d (%.1f%%) | threads:%d | bw:%.1f Mbps",
                        completed, total_rows,
                        (completed / total_rows) * 100,
                        n_threads, bw_monitor.current_mbps,
                    )

    except KeyboardInterrupt:
        overall.close()
        log.warning("")
        log.warning("=" * 60)
        log.warning("INTERRUPTED — download stopped by user (Ctrl+C)")
        log.warning("Progress saved — resume by running the same command again.")
        log.warning("=" * 60)
        log.info("Partial results — Downloaded:%d Skipped:%d Failed:%d",
                 counts["ok"], counts["skipped"], counts["failed"])
        sys.exit(0)

    overall.close()

    # --- Step 10: Final summary ----------------------------------------------

    log.info("=" * 60)
    log.info("DOWNLOAD COMPLETE — SUMMARY")
    log.info("=" * 60)
    log.info("  Total      : %d", len(rows))
    log.info("  Downloaded : %d", counts["ok"])
    log.info("  Skipped    : %d", counts["skipped"])
    log.info("  Failed     : %d", counts["failed"])
    log.info("  No URL     : %d", counts["no_url"])
    log.info("-" * 60)
    log.info("  Output     : %s", output_dir.resolve())
    log.info("  Log file   : %s", log_file)
    if counts["failed"] > 0:
        log.info("  Failed log : %s", str(Path(LOG_DIR) / "failed.log"))
        log.info("  To retry   : python download_videos.py --retry-failed")
    log.info("  Progress   : %s", PROGRESS_FILE)
    log.info("  Final BW   : %.1f Mbps | Final threads: %d", bw_monitor.current_mbps, bw_monitor.current_threads)
    log.info("=" * 60)

if __name__ == "__main__":
    main()